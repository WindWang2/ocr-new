import re
"""
OCR 实验 API 服务

功能：
1. 相机管理（增删改查）
2. 实验执行（触发单字段相机拍照并保存读数）
3. 实验记录查询

启动: uvicorn main:app --reload --port 8001
"""

from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime
import logging
import json
import re
import io
import os
import openpyxl
import asyncio
import time
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys

# 添加项目根目录到路径
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from backend.models.database import (
    init_db, add_camera, get_cameras, get_camera_by_id,
    create_experiment,
    get_experiment, list_experiments, delete_experiment,
    create_reading, upsert_reading, get_readings_by_experiment, get_connection,
    get_config, set_config,
    get_all_templates, get_template, upsert_template,
)
from backend.services.camera_control import (
    CameraClient, get_all_enabled_cameras
)
from backend.services.mock_camera import MockCameraClient
from backend.services.task_manager import task_manager, TaskState
from backend.services.multi_instrument_pipeline import MultiInstrumentPipeline
from config import Config

# 日志配置
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s"
)
logger = logging.getLogger(__name__)

# 初始化数据库
init_db()

# FastAPI 应用
app = FastAPI(
    title="OCR 实验服务 API",
    description="多相机仪表读数识别实验管理",
    version="1.0.0"
)

# 全局 OCR 流水线 (懒加载)
_pipeline = None

def get_pipeline():
    global _pipeline
    if _pipeline is None:
        logger.info("正在初始化 MultiInstrumentPipeline...")
        _pipeline = MultiInstrumentPipeline()
    return _pipeline

# CORS 中间件
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    import traceback
    error_msg = traceback.format_exc()
    logger.error(f"Unhandled exception at {request.url}: {error_msg}")
    return JSONResponse(
        status_code=500,
        content={"success": False, "detail": "Internal Server Error", "traceback": error_msg},
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "*",
            "Access-Control-Allow-Headers": "*",
        }
    )

from backend.services.path_utils import normalize_path

# 静态文件：智能挂载图片目录
# 优先从数据库读取，如果没有则使用 Config 默认值
image_dir_path = get_config("image_dir") or Config.CAMERA_IMAGE_DIR
_images_dir = normalize_path(image_dir_path)
if not _images_dir.is_absolute():
    # 尝试基于项目根目录定位
    _images_dir = Path(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", str(_images_dir))))

_images_dir.mkdir(parents=True, exist_ok=True)
logger.info(f"图片映射根目录: {_images_dir}")

from fastapi.responses import FileResponse

@app.on_event("startup")
async def startup_event():
    # 启动异步清理任务
    asyncio.create_all_tasks = False # Compatibility
    # asyncio.create_task(image_cleanup_task())
    logger.info("后台清理任务已启动")

@app.get("/images/{path:path}")
async def serve_image(path: str):
    """
    智能提供图片文件服务：
    1. 寻找原始文件。
    2. 如果是 .jpg 但文件不存在，尝试寻找对应的 .png (针对新无损流程的平滑过渡)。
    """
    file_path = _images_dir / path
    
    # 1. 尝试原始路径
    if file_path.exists() and file_path.is_file():
        return FileResponse(str(file_path))
    
    # 2. .jpg -> .png 自动容错逻辑
    if str(path).lower().endswith(".jpg"):
        png_path = _images_dir / (str(path)[:-4] + ".png")
        if png_path.exists():
            return FileResponse(str(png_path))

    # 3. 如果还是没有，尝试在 crops 及其子目录下寻找 (智能纠错)
    if "crops" not in str(path):
        p = Path(path)
        # 优先尝试新结构：parent/crops/display/name
        alt_path_display = _images_dir / p.parent / "crops" / "display" / p.name
        if alt_path_display.exists(): return FileResponse(str(alt_path_display))
            
        # 尝试：parent/crops/name (旧结构)
        alt_path = _images_dir / p.parent / "crops" / p.name
        if alt_path.exists(): return FileResponse(str(alt_path))
            
        # 同时也检查 png 格式 (如果是旧的 jpg 引用)
        alt_path_png = _images_dir / p.parent / "crops" / (p.stem + ".png")
        if alt_path_png.exists(): return FileResponse(str(alt_path_png))
            
        # 检查 display 下的 png
        alt_path_display_png = _images_dir / p.parent / "crops" / "display" / (p.stem + ".png")
        if alt_path_display_png.exists(): return FileResponse(str(alt_path_display_png))

    raise HTTPException(status_code=404, detail=f"图片不存在: {path}")


async def image_cleanup_task():
    """定期清理过期的图片文件，防止磁盘占满"""
    retention_days = float(os.getenv("IMAGE_RETENTION_DAYS", "7"))
    while True:
        try:
            now = time.time()
            for root, dirs, files in os.walk(str(_images_dir)):
                for file in files:
                    # 安全检查：仅删除图片文件，且跳过隐藏目录（如 .git）
                    if not file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp')):
                        continue
                    if '.git' in root or '.gemini' in root:
                        continue
                        
                    file_path = os.path.join(root, file)
                    if now - os.path.getmtime(file_path) > retention_days * 86400:
                        os.remove(file_path)
                        logger.info(f"Cleaned up old image: {file_path}")
        except Exception as e:
            logger.error(f"图片清理任务出错: {e}")
        await asyncio.sleep(86400)  # 每天运行一次

@app.on_event("startup")
async def startup_event():
    # 同步数据库配置到全局 Config
    db_image_dir = get_config("image_dir")
    if db_image_dir:
        Config.update_image_dir(db_image_dir)
        logger.info(f"已同步数据库图片目录到 Config: {db_image_dir}")

    # asyncio.create_task(image_cleanup_task())
    # asyncio.create_task(_task_cleanup_loop())


async def _task_cleanup_loop():
    """定期清理过期的异步任务记录"""
    while True:
        try:
            task_manager.cleanup()
        except Exception as e:
            logger.error(f"任务清理出错: {e}")
        await asyncio.sleep(300)  # 每5分钟清理一次


# ==================== 异步任务查询 API ====================

@app.get("/tasks/{task_id}")
def get_task_status(task_id: str):
    """查询异步任务状态和结果"""
    info = task_manager.get_status(task_id)
    if not info:
        raise HTTPException(status_code=404, detail="任务不存在")
    resp = {
        "task_id": info.task_id,
        "status": info.status.value,
        "progress": info.progress,
        "message": info.message,
        "created_at": info.created_at,
    }
    if info.started_at:
        resp["started_at"] = info.started_at
    if info.completed_at:
        resp["completed_at"] = info.completed_at
    if info.status == TaskState.COMPLETED and info.result is not None:
        resp["result"] = info.result
    if info.status == TaskState.FAILED and info.error:
        resp["error"] = info.error
    return resp


def _convert_and_save_image(raw_path: str, camera_id: int, run_index: int) -> Optional[str]:
    """
    将相机拍摄的图片（可能是BMP）转换为JPG格式，缩放最长边到500px，保存到 camera_images/ 目录。
    返回相对于 camera_images/ 的路径（如 F0/001.jpg），供前端通过 /images/ 访问。
    """
    from PIL import Image

    src = Path(raw_path)
    if not src.exists():
        return None

    # 完全禁用预览图生成逻辑
    # 获取相对于 _images_dir 的相对路径，供前端访问
    import os
    try:
        abs_path = os.path.abspath(raw_path)
        rel_path = os.path.relpath(abs_path, _images_dir)
        return rel_path.replace('\\', '/')
    except:
        return raw_path


# ==================== 请求模型 ====================

class TemplateField(BaseModel):
    name: str
    label: str
    unit: Optional[str] = ""
    type: str = "float"

class InstrumentTemplateCreate(BaseModel):
    instrument_type: str
    name: str
    description: Optional[str] = ""
    prompt_template: str
    fields: List[TemplateField]
    keywords: List[str]
    example_images: Optional[List[str]] = []
    default_tier: int = 2

class CameraCreate(BaseModel):
    name: str
    camera_id: int
    control_host: Optional[str] = "127.0.0.1"
    control_port: Optional[int] = None
    mode: Optional[str] = 'single'


class CameraConfigItem(BaseModel):
    field_key: str
    camera_id: int
    max_readings: int
    selected_readings: Optional[List[str]] = None


class ExperimentCreate(BaseModel):
    name: str
    type: str  # kinematic_viscosity | apparent_viscosity | surface_tension
    manual_params: Optional[dict] = {}
    camera_configs: Optional[List[CameraConfigItem]] = []
    description: Optional[str] = None


class ExperimentRunField(BaseModel):
    field_key: str
    camera_id: int


class ExperimentCaptureBody(BaseModel):
    camera_id: int
    target_instrument_id: Optional[int] = None
    field_key: Optional[str] = None


class InstrumentMappingUpdate(BaseModel):
    mapping: Dict[str, int]


# ==================== 相机管理 API ====================

@app.get("/cameras")
def list_cameras(enabled_only: bool = True):
    """获取相机列表"""
    cameras = get_cameras(enabled_only=enabled_only)
    return {"success": True, "count": len(cameras), "cameras": cameras}


@app.get("/cameras/{camera_id}")
def get_camera(camera_id: int):
    """获取单个相机信息"""
    camera = get_camera_by_id(camera_id)
    if not camera:
        raise HTTPException(status_code=404, detail="相机不存在")
    return {"success": True, "camera": camera}


@app.post("/cameras")
def create_camera(camera: CameraCreate):
    """添加新相机"""
    try:
        camera_id_db = add_camera(
            name=camera.name,
            camera_id=camera.camera_id,
            control_host=camera.control_host,
            control_port=camera.control_port,
            mode=camera.mode or 'single',
        )
        return {"success": True, "camera_id": camera_id_db, "message": "相机添加成功"}
    except Exception as e:
        logger.error(f"添加相机失败: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/cameras/match_instruments")
async def match_instruments():
    """
    一键匹配仪器到相机：
    1. 依次触发所有可用相机拍照
    2. 利用 YOLO 识别各相机中的仪器
    3. 更新系统配置中的 instrument_camera_mapping
    """
    logger.info("开始一键仪器匹配...")
    enabled_cameras = get_cameras(enabled_only=True)
    if not enabled_cameras:
        return {"success": False, "detail": "没有已启用的相机"}

    pipeline = get_pipeline()
    
    # instrument_id -> {camera_id, confidence}
    best_matches = {}
    scan_results = []

    for cam in enabled_cameras:
        cam_id = cam["camera_id"]
        logger.info(f"正在扫描相机 {cam_id}...")
        
        # 拍照
        mock_enabled = get_config("mock_camera_enabled", default=False)
        logger.info(f"相机 {cam_id}: Mock模式={mock_enabled}")
        try:
            image_dir = get_config("image_dir", default=None) or None
            if mock_enabled:
                logger.info(f"相机 {cam_id}: 正在从目录 {image_dir} 获取 Mock 图片")
                client = MockCameraClient(camera_id=cam_id, image_dir=image_dir)
            else:
                camera_config = Config.get_camera_config()
                if image_dir:
                    camera_config["image_dir"] = image_dir
                client = CameraClient(camera_id=cam_id, config=camera_config)
            
            success, capture_result = client.capture_image()
            if not success:
                scan_results.append({"camera_id": cam_id, "status": "capture_failed", "error": capture_result.get("error")})
                continue
            
            image_path = capture_result.get("image_path")
            if not image_path:
                logger.warning(f"相机 {cam_id}: 未获取到图片路径")
                scan_results.append({"camera_id": cam_id, "status": "no_image"})
                continue
            
            logger.info(f"相机 {cam_id}: 成功获取图片 {image_path}")

            # 转换为 PIL 图片进行处理
            from PIL import Image
            img = Image.open(image_path).convert("RGB")
            
            # YOLO 检测
            detections = pipeline.yolo_detector.detect(img)
            found_instruments = []
            
            logger.info(f"相机 {cam_id}: 原始检测到 {len(detections)} 个目标")
            for det in detections:
                conf = float(det[4])
                class_id = int(det[5])
                logger.info(f"  - 目标 ID: {class_id}, 置信度: {conf:.4f}")
                
                # 严格限制标签范围 0-8
                if not (0 <= class_id <= 8):
                    logger.warning(f"    [跳过] ID {class_id} 不在 0-8 范围内")
                    continue
                    
                found_instruments.append({"class_id": class_id, "confidence": conf})
                
                # 如果这是该仪器目前遇到的最高置信度，则记录下来
                if class_id not in best_matches or conf > best_matches[class_id]["confidence"]:
                    best_matches[class_id] = {
                        "camera_id": cam_id,
                        "confidence": conf
                    }
            
            scan_results.append({
                "camera_id": cam_id,
                "status": "success",
                "instruments": found_instruments
            })
            
        except Exception as e:
            logger.error(f"扫描相机 {cam_id} 出错: {e}")
            scan_results.append({"camera_id": cam_id, "status": "error", "detail": str(e)})

    # 更新全局映射
    if best_matches:
        new_mapping = {str(k): v["camera_id"] for k, v in best_matches.items()}
        set_config("instrument_camera_mapping", new_mapping)
        logger.info(f"一键匹配完成，新映射: {new_mapping}")
        
        # 返回结果中包含名称
        from instrument_reader import DynamicInstrumentLibrary
        final_summary = []
        for inst_id, data in best_matches.items():
            name = DynamicInstrumentLibrary.get_instrument_type_from_camera(f"F{inst_id}")
            final_summary.append({
                "instrument_id": inst_id,
                "instrument_name": name,
                "camera_id": data["camera_id"],
                "confidence": data["confidence"]
            })
        
        return {
            "success": True,
            "mapping": new_mapping,
            "summary": final_summary,
            "scan_details": scan_results
        }
    else:
        return {
            "success": False,
            "detail": "扫描完成，但未识别到任何已知仪表",
            "scan_details": scan_results
        }


@app.delete("/cameras/{camera_id}")
def remove_camera(camera_id: int):
    """删除相机（软删除设置为 disabled）"""
    # 这里使用软删除，实际禁用
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE cameras SET enabled = 0 WHERE camera_id = ?", (camera_id,))
    conn.commit()
    deleted = cursor.rowcount > 0
    conn.close()

    if not deleted:
        raise HTTPException(status_code=404, detail="相机不存在")
    return {"success": True, "message": "相机已禁用"}


# ==================== 实验管理 API ====================

@app.get("/experiments")
def list_experiments_api(
    limit: int = Query(50, le=100),
    offset: int = Query(0, ge=0),
):
    """获取实验列表（摘要，不含读数）"""
    experiments = list_experiments(limit=limit, offset=offset)
    return {"success": True, "count": len(experiments), "experiments": experiments}


@app.get("/experiments/{exp_id}")
def get_experiment_api(exp_id: int):
    """获取实验详情（含所有读数）"""
    experiment = get_experiment(exp_id)
    if not experiment:
        raise HTTPException(status_code=404, detail="实验不存在")
    return {"success": True, "experiment": experiment}


@app.post("/experiments")
def create_experiment_api(exp: ExperimentCreate):
    """创建实验记录"""
    VALID_TYPES = {"kinematic_viscosity", "apparent_viscosity", "surface_tension", "water_mineralization", "ph_value", "test"}
    if exp.type not in VALID_TYPES:
        raise HTTPException(status_code=400, detail=f"无效实验类型: {exp.type}")
    try:
        exp_id = create_experiment(
            name=exp.name,
            exp_type=exp.type,
            manual_params=exp.manual_params,
            camera_configs=[c.dict() for c in exp.camera_configs],
            description=exp.description,
        )
        return {"success": True, "experiment_id": exp_id}
    except Exception as e:
        logger.error(f"创建实验失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/experiments/{exp_id}/run")
def run_experiment_field(exp_id: int, body: ExperimentRunField):
    """
    触发单个字段的相机拍照→OCR→保存读数 (已升级为现代 YOLO+LLM 流程)
    """
    result = _core_run_test_capture(
        exp_id=exp_id,
        field_key=body.field_key,
        camera_id=body.camera_id,
        run_index=body.run_index
    )
    if not result.get("success"):
        return {"success": False, "detail": result.get("detail", "识别失败")}
        
    return {"success": True, "reading": result.get("reading")}


@app.post("/experiments/{exp_id}/capture")
def capture_image_endpoint(exp_id: int, body: ExperimentCaptureBody):
    """
    仅拍照并保存图片（不做OCR），立即返回图片路径供前端显示。
    前端收到图片后可立即展示，再调用 /run-test 进行 OCR。
    """
    experiment = get_experiment(exp_id)
    if not experiment:
        raise HTTPException(status_code=404, detail="实验不存在")

    mock_enabled = get_config("mock_camera_enabled", default=False)
    image_dir_config = get_config("image_dir", default=None) or None

    try:
        if mock_enabled:
            client = MockCameraClient(camera_id=body.camera_id, image_dir=image_dir_config)
            success, result = client.capture_image()
        else:
            camera_config = Config.get_camera_config()
            if image_dir_config:
                camera_config["image_dir"] = image_dir_config
            client = CameraClient(camera_id=body.camera_id, config=camera_config)
            success, result = client.trigger_and_read()
    except Exception as e:
        logger.error(f"相机 {body.camera_id} 拍照失败: {e}")
        return {"success": False, "detail": f"相机连接失败: {e}"}

    if not success:
        return {"success": False, "detail": result.get("error", "拍照失败")}

    # 获取原始高分辨率路径（BMP）
    original_path = result.get("raw_image_path") or result.get("image_path")
    if not original_path:
        return {"success": False, "detail": "未获取到图片"}

    # ── 直接用原图走 YOLO，不再生成 500px 预览 ──────────────────────
    try:
        from instrument_reader import InstrumentReader
        reader = InstrumentReader()
        raw_abs_path = os.path.abspath(str(original_path))
        if not os.path.exists(raw_abs_path):
            return {"success": False, "detail": f"原始图片不存在: {raw_abs_path}"}

        detect_result = reader.detect_only(raw_abs_path, target_class_id=body.target_instrument_id)
        if detect_result["success"] and detect_result["results"]:
            detected_ids = [r["class_id"] for r in detect_result["results"]]
            logger.info(f"Capture YOLO 检测到: {detected_ids}")

            # 如果指定了目标仪器，找对应的 crop
            if body.target_instrument_id is not None:
                best_crop = next((r for r in detect_result["results"] if r["class_id"] == body.target_instrument_id), None)
                if best_crop:
                    logger.info(f"Capture 对位成功: D{body.target_instrument_id} -> {best_crop['cropped_image_path']}")
                    return {"success": True, "image_path": best_crop["cropped_image_path"], "camera_id": body.camera_id}

            # 未指定目标或未匹配到目标，返回第一个检测结果
            first_crop = detect_result["results"][0]
            return {"success": True, "image_path": first_crop["cropped_image_path"], "camera_id": body.camera_id}

    except Exception as e:
        import traceback
        logger.warning(f"Capture YOLO 检测失败: {e}\n{traceback.format_exc()}")

    # ── YOLO 失败兜底：保存一张 600px 的全图 ────────────────────────
    from PIL import Image as _PILImage
    try:
        img = _PILImage.open(raw_abs_path).convert("RGB")
        w, h = img.size
        max_side = 600
        if max(w, h) > max_side:
            s = max_side / max(w, h)
            img = img.resize((int(w * s), int(h * s)), _PILImage.Resampling.LANCZOS)
        fallback_dir = Path(raw_abs_path).parent / "crops"
        fallback_dir.mkdir(parents=True, exist_ok=True)
        fallback_name = f"fallback_{body.camera_id}_{__import__('time').strftime('%H%M%S')}.png"
        fallback_path = fallback_dir / fallback_name
        img.save(str(fallback_path), "PNG")
        # 计算相对路径
        parts = list(fallback_path.parts)
        start_idx = 0
        for i in range(len(parts) - 1, -1, -1):
            if re.match(r'^[Ff]\d+$', parts[i]):
                start_idx = i
                break
        rel = "/".join(parts[start_idx:])
        return {"success": True, "image_path": rel.replace("\\", "/"), "camera_id": body.camera_id}
    except Exception as e2:
        logger.error(f"Fallback 也失败: {e2}")
        return {"success": False, "detail": f"YOLO 和 Fallback 均失败"}

@app.post("/test/detect_only")
def test_detect_only(path: str):
    from instrument_reader import InstrumentReader
    reader = InstrumentReader()
    return reader.detect_only(path)

class ExperimentRunTestBody(BaseModel):
    field_key: str
    camera_id: Optional[int] = None      # 物理相机 ID (可选，不传则根据仪器自动解析)
    target_instrument_id: Optional[int] = None # 目标仪器类别 ID (0-8, 用于 YOLO 过滤)
    image_path: Optional[str] = None
    reading_key: Optional[str] = None   # 仪器读数键，如 "actual_reading"、"tension"
    run_index: Optional[int] = None     # 槽位序号（0-based），不传则追加
    precise: bool = False               # 精准识别：先 OCR 提取文字识别仪器类型，再二次读数
    camera_mode: Optional[str] = None  # D0 专用："auto"（自动模式）或 "manual"（手动模式）


def _core_run_test_capture(
    exp_id: int,
    field_key: str,
    camera_id: Optional[int] = None,
    target_instrument_id: Optional[int] = None,
    image_path: Optional[str] = None,
    reading_key: Optional[str] = None,
    run_index: Optional[int] = None,
    precise: bool = False,
    camera_mode: Optional[str] = None,
) -> dict:
    """
    核心 OCR 测试逻辑（同步/异步版共享）。

    完成以下步骤:
    1. 解析目标仪器 ID 并路由到物理相机
    2. 拍照（若未提供 image_path）
    3. YOLO + LLM OCR 识别
    4. 目标匹配 + 白名单过滤
    5. 主值提取与保存

    返回 dict（不抛出 HTTPException），由上层端点决定异常策略。
    """
    from backend.models.constants import INSTRUMENT_FIELD_WHITELIST as WHITELIST
    import os

    experiment = get_experiment(exp_id)
    if not experiment:
        return {"success": False, "detail": "实验不存在"}

    # 1. 确定目标仪器 ID
    if target_instrument_id is None:
        match = re.search(r'D(\d+)', field_key)
        if match:
            target_instrument_id = int(match.group(1))
        else:
            # 如果 field_key 也不包含 Fx，则尝试回退到 camera_id
            target_instrument_id = camera_id if camera_id is not None else 0

    # 自动根据仪器 ID 路由到物理相机
    from instrument_reader import DynamicInstrumentLibrary
    physical_camera_id = DynamicInstrumentLibrary.get_physical_camera_id(target_instrument_id)
    
    # 如果明确传了 camera_id，则以传入的为准（用于强制指定相机测试）
    if camera_id is not None:
        physical_camera_id = camera_id

    saved_image_path = image_path
    high_res_src_path = None # 初始化

    if not saved_image_path:
        mock_enabled = get_config("mock_camera_enabled", default=False)
        try:
            image_dir = get_config("image_dir", default=None) or None
            if mock_enabled:
                from backend.services.mock_camera import MockCameraClient
                client = MockCameraClient(camera_id=physical_camera_id, image_dir=image_dir)
                success, result = client.capture_image()
            else:
                from backend.services.camera_control import CameraClient
                camera_config = Config.get_camera_config()
                if image_dir:
                    camera_config["image_dir"] = image_dir
                client = CameraClient(camera_id=physical_camera_id, config=camera_config)
                success, result = client.capture_image()
        except Exception as e:
            logger.error(f"物理相机 {physical_camera_id} (仪器 D{target_instrument_id}) 拍照失败: {e}")
            return {"success": False, "detail": f"相机连接失败: {e}"}

        if not success:
            return {"success": False, "detail": result.get("error", "拍照失败")}

        preview_src_path = result.get("image_path")
        high_res_src_path = result.get("raw_image_path") # 优先使用原始 BMP/JPG
        
        # 核心修复：不再调用 _convert_and_save_image，因为它会用缩略图覆盖同名的原图。
        # 直接使用 CameraClient 已经处理好的路径。
        if preview_src_path:
            # 获取相对于 _images_dir 的相对路径，供前端访问
            try:
                saved_image_path = os.path.relpath(preview_src_path, _images_dir)
            except:
                saved_image_path = preview_src_path

    if not saved_image_path:
        return {"success": False, "detail": "无可用图片"}

    # OCR
    # 核心修复：优先在原始高分辨率图上执行 OCR (如果存在)
    # 如果 saved_image_path 本身就是绝对路径且存在，直接用它
    if os.path.isabs(str(saved_image_path)) and os.path.exists(str(saved_image_path)):
        ocr_input_path = str(saved_image_path)
    else:
        ocr_input_path = high_res_src_path if (high_res_src_path and os.path.exists(high_res_src_path)) else os.path.abspath(str(_images_dir / saved_image_path))
    
    logger.info(f"==> OCR 输入路径: {ocr_input_path}")
    print(f"\n[MAIN_DEBUG] target_instrument_id={target_instrument_id}, physical_camera_id={physical_camera_id}, saved_image_path={saved_image_path}\n")
    try:
        from instrument_reader import InstrumentReader
        from backend.services.llm_provider import get_global_provider
        reader = InstrumentReader(provider=get_global_provider())
        ocr_start = time.time()
        # 优化：如果是裁剪后的图片且已知仪器 ID，直接走专用 Prompt 通道，跳过冗余 YOLO
        if target_instrument_id is not None and "crop" in str(saved_image_path):
            logger.info(f"检测到裁剪图，直接执行 D{target_instrument_id} 专用识别...")
            ocr_result = reader.read_instrument(ocr_input_path, target_class_id=target_instrument_id)
        else:
            ocr_result = reader.read_instrument(ocr_input_path, target_class_id=target_instrument_id)
        ocr_ms = (time.time() - ocr_start) * 1000
    except Exception as e:
        logger.error(f"OCR 失败: {e}")
        return {"success": False, "detail": f"OCR 识别失败: {e}"}

    if not ocr_result.get("success"):
        return {
            "success": False, 
            "detail": ocr_result.get("error", "OCR 识别失败"),
            "image_path": ocr_result.get("cropped_image_path", saved_image_path)
        }

    # 2. 精准匹配目标仪器
    best_target = None
    all_results = ocr_result.get("all_results", []) if ocr_result.get("multiple_targets") else [ocr_result]

    if all_results:
        # 2a. 查精确匹配
        best_target = next((t for t in all_results if t.get("class_id") == target_instrument_id), None)
        
        # 2b. 共享相机模糊匹配
        if not best_target and target_instrument_id is not None:
            from instrument_reader import DynamicInstrumentLibrary
            all_ids_on_this_camera = sorted([id for id in range(9) if DynamicInstrumentLibrary.get_physical_camera_id(id) == physical_camera_id])
            if target_instrument_id in all_ids_on_this_camera:
                target_idx = all_ids_on_this_camera.index(target_instrument_id)
                sorted_res = sorted(all_results, key=lambda x: (x.get("bbox") or [0])[0])
                if target_idx < len(sorted_res):
                    best_target = sorted_res[target_idx]
                    logger.info(f"F{target_instrument_id} 精确匹配失败，采用位置排序匹配(OCR阶段): Index {target_idx}")

    if not best_target:
        return {"success": True, "detail": f"图中未检测到目标仪表 (ID:{target_instrument_id})", "all_ocr": {}, "readings": [], "image_path": saved_image_path}

    # 3. 解析读数
    raw_response = best_target.get("readings", {})
    readings_dict = raw_response if isinstance(raw_response, dict) else {}
    if isinstance(raw_response, str):
        try:
            readings_dict = json.loads(raw_response)
        except Exception:
            readings_dict = {}

    # 4. 白名单过滤与数值提取 (增强版：双向键名对齐)
    from instrument_reader import DynamicInstrumentLibrary
    template = DynamicInstrumentLibrary.get_template(str(target_instrument_id))
    fields_map = {} # 英文键 -> 中文标签
    reverse_fields_map = {} # 中文标签 -> 英文键
    if template:
        for f in template.get('fields', []):
            fields_map[f['name']] = f['label']
            reverse_fields_map[f['label']] = f['name']

    current_whitelist = WHITELIST.get(target_instrument_id, [])
    filtered_readings = {}
    numeric_ocr = {}

    for k, v in readings_dict.items():
        if v is None: continue
        
        # 兼容性修复：如果大模型按要求输出了中文键名，自动转为英文内部键
        if k in reverse_fields_map:
            k = reverse_fields_map[k]
            
        # 字段过滤 (使用英文键过滤)
        if target_instrument_id is not None and k not in current_whitelist:
            continue

        # 存入原始值 (英文键)
        filtered_readings[k] = v
        # 如果有对应的中文标签，映射一份到 filtered_readings 和 numeric_ocr (双向填充)
        label = fields_map.get(k)
        if label:
            filtered_readings[label] = v
        
        # 智能提取数值用于计算主读数
        try:
            val = None
            if isinstance(v, (int, float)):
                val = float(v)
            else:
                str_val = str(v).strip()
                if ":" in str_val and k == "time": # 时间转换
                    parts = str_val.split(":")
                    val = float(parts[0]) * 60 + float(parts[1])
                else:
                    clean_val = re.sub(r'[^\d\.\-]', '', str_val)
                    if clean_val and clean_val != ".": val = float(clean_val)
            
            if val is not None:
                numeric_ocr[k] = val
                if label:
                    numeric_ocr[label] = val # 核心：将中文标签也加入数值字典
        except: pass

    # 5. 确定主值并保存
    primary_key = reading_key
    if not primary_key:
        config = next((c for c in experiment["camera_configs"] if c["field_key"] == field_key), None)
        selected = config.get("selected_readings", []) if config else []
        # 在识别出的数字里找第一个被选中的字段 (支持中英文)
        primary_key = next((k for k in selected if k in numeric_ocr), None)
        
        if not primary_key:
            # 智能兜底
            if target_instrument_id in (1, 2): primary_key = 'weight'
            elif target_instrument_id == 3: primary_key = 'ph_value'
            elif target_instrument_id == 5: primary_key = 'tension'
            elif target_instrument_id == 8: primary_key = 'actual_reading'
            elif numeric_ocr: primary_key = list(numeric_ocr.keys())[0]

    primary_value = numeric_ocr.get(primary_key, 0.0) if primary_key else 0.0
    
    logger.info(f"D{target_instrument_id} 识别结果汇总: {numeric_ocr}, 选定主值: {primary_key}={primary_value}")

    target_image_path = best_target.get("cropped_image_path", saved_image_path)
    final_run_index = run_index if run_index is not None else len(
        [r for r in experiment["readings"] if r["field_key"] == field_key]
    )

    # 准备 OCR 数据，包含识别图路径和检测框以便审计
    ocr_metadata = filtered_readings.copy()
    ocr_metadata["_full_image_path"] = saved_image_path
    
    # 性能指标
    if 'ocr_ms' in locals():
        ocr_metadata["_performance"] = {
            "total_ms": round(ocr_ms, 1)
        }
    
    # 提取并保存原始 LLM 输出 (如果存在)
    if "_raw_output" in readings_dict:
        ocr_metadata["_raw_output"] = readings_dict["_raw_output"]
        
    if best_target.get("recognition_image_path"):
        ocr_metadata["_recognition_image_path"] = best_target["recognition_image_path"]
    if best_target.get("bbox"):
        ocr_metadata["_bbox"] = best_target["bbox"]

    reading = upsert_reading(
        experiment_id=exp_id,
        field_key=field_key,
        camera_id=physical_camera_id,
        value=primary_value,
        run_index=final_run_index,
        image_path=target_image_path,
        ocr_data=ocr_metadata # 记录包含识别图路径的完整元数据
    )

    return {
        "success": True,
        "readings": [reading],
        "all_ocr": filtered_readings,
        "image_path": target_image_path, # 核心：将裁剪图路径传回前端，确保 UI 同步
        "detail": None if (primary_key and primary_key in numeric_ocr) else "已精准捕获仪器特写，但读数需校对",
    }

@app.post("/experiments/{exp_id}/detect")
def detect_instrument_endpoint(exp_id: int, body: ExperimentRunTestBody):
    """
    仅运行 YOLO 检测并回传裁剪后的图片路径 (不进行 LLM 识别)。
    如果未提供 image_path，会自动解析物理相机路径。
    """
    import os
    from instrument_reader import InstrumentReader, DynamicInstrumentLibrary

    # 1. 确定物理路径
    image_path = body.image_path
    if not image_path:
        return {"success": False, "detail": "必须提供 image_path"}

    full_image_path = os.path.abspath(str(_images_dir / image_path))
    
    try:
        reader = InstrumentReader()
        result = reader.detect_only(full_image_path)
        
        if result["success"]:
            # 找到匹配 target_instrument_id 的结果
            target_id = body.target_instrument_id
            if target_id is None:
                match = re.search(r'F(\d+)', body.field_key)
                if match: target_id = int(match.group(1))

            best_crop = None
            if target_id is not None:
                best_crop = next((r for r in result["results"] if r["class_id"] == target_id), None)

            if best_crop:
                return {
                    "success": True,
                    "image_path": best_crop["cropped_image_path"],
                    "class_id": best_crop["class_id"],
                    "bbox": best_crop["bbox"]
                }
        
        return {"success": False, "detail": result.get("error", "未检出结果"), "image_path": image_path}
    except Exception as e:
        logger.error(f"检测失败: {e}")
        return {"success": False, "detail": str(e), "image_path": image_path}


@app.post("/experiments/{exp_id}/auto-trigger")
def auto_trigger_instrument(exp_id: int, body: ExperimentRunTestBody):
    """
    自动触发: 
    1. 根据 instrument_id (0-8) 自动寻找物理相机
    2. 拍照
    3. YOLO 检测目标
    4. 裁剪并存图
    5. 返回 crop 路径，供前端立即显示
    """
    import os
    from instrument_reader import InstrumentReader, DynamicInstrumentLibrary
    
    # 1. 确定目标仪器 ID (F0-F8)
    target_id = body.target_instrument_id
    if target_id is None:
        match = re.search(r'F(\d+)', body.field_key)
        if match: 
            target_id = int(match.group(1))
    
    if target_id is None:
        # 如果还是没找到，尝试从 camera_id 映射 (旧逻辑兼容)
        target_id = body.camera_id

    # 2. 自动定位物理相机
    physical_camera_id = DynamicInstrumentLibrary.get_physical_camera_id(target_id)
    logger.info(f"仪器 D{target_id} 映射到物理相机 {physical_camera_id}")

    # 3. 拍照
    mock_enabled = get_config("mock_camera_enabled", default=False)
    image_dir_cfg = get_config("image_dir", default=None) or None
    try:
        if mock_enabled:
            client = MockCameraClient(camera_id=physical_camera_id, image_dir=image_dir_cfg)
            success, result = client.capture_image()
        else:
            camera_config = Config.get_camera_config()
            if image_dir_cfg:
                camera_config["image_dir"] = image_dir_cfg
            client = CameraClient(camera_id=physical_camera_id, config=camera_config)
            success, result = client.capture_image()
    except Exception as e:
        logger.error(f"拍照失败: {e}")
        return {"success": False, "detail": f"相机连接失败: {e}"}

    if not success:
        return {"success": False, "detail": result.get("error", "拍照失败")}

    # 4. 转换并保存全图 (用于前端预览)
    # 注意：raw_image_path 可能已经是预览图，取决于 CameraClient 的返回
    preview_image_path = result.get("image_path")
    original_image_path = result.get("raw_image_path")
    
    # 生成用于前端显示的预览图路径（相对于 images 目录）
    saved_full_path = _convert_and_save_image(original_image_path or preview_image_path, physical_camera_id, 0)
    
    # 5. YOLO 检测并获取特写
    # 【核心修复】必须使用原始高分辨率图进行检测和裁剪，以保证特写图的清晰度
    detection_input = original_image_path if (original_image_path and os.path.exists(original_image_path)) else os.path.abspath(str(_images_dir / saved_full_path))
    
    logger.info(f"YOLO 检测输入: {detection_input}")
    try:
        reader = InstrumentReader()
        detect_result = reader.detect_only(detection_input)
        
        if detect_result["success"] and detect_result["results"]:
            results = detect_result["results"]
            logger.info(f"Target D{target_id} detection results: {[r['class_id'] for r in results]}")
            
            # 严格匹配 class_id == target_id
            best_crop = next((r for r in results if r["class_id"] == target_id), None)
            
            if best_crop:
                return {
                    "success": True,
                    "image_path": saved_full_path, 
                    "cropped_image_path": best_crop["cropped_image_path"], 
                    "class_id": best_crop["class_id"],
                    "bbox": best_crop["bbox"],
                    "camera_id": physical_camera_id
                }
        
        # 彻底失败回退全图
        logger.warning(f"F{target_id} 目标定位失败，返回全景图。全图路径: {saved_full_path}")
        return {
            "success": True, 
            "image_path": saved_full_path, 
            "cropped_image_path": saved_full_path, 
            "detail": "未检出目标，回退全图",
            "camera_id": physical_camera_id
        }
    except Exception as e:
        logger.error(f"检测出错: {e}")
        return {
            "success": True, 
            "image_path": saved_full_path, 
            "cropped_image_path": saved_full_path, 
            "detail": str(e),
            "camera_id": physical_camera_id
        }


@app.post("/experiments/{exp_id}/run-test")
def run_test_capture(exp_id: int, body: ExperimentRunTestBody):
    """
    测试模板 OCR 识别。
    如果提供 image_path 则跳过拍照，直接对指定图片做 OCR。
    """
    experiment = get_experiment(exp_id)
    if not experiment:
        raise HTTPException(status_code=404, detail="实验不存在")

    return _core_run_test_capture(
        exp_id=exp_id,
        field_key=body.field_key,
        camera_id=body.camera_id,
        target_instrument_id=body.target_instrument_id,
        image_path=body.image_path,
        reading_key=body.reading_key,
        run_index=body.run_index,
        precise=body.precise,
        camera_mode=body.camera_mode,
    )


# ==================== 异步任务版本 API ====================
# 以下端点将耗时操作（拍照 + OCR + LLM）放入后台线程执行，
# 立即返回 task_id，前端通过 GET /tasks/{task_id} 轮询结果。

def _do_run_experiment_field(exp_id: int, field_key: str, camera_id: int):
    """后台线程：执行 run_experiment_field 的实际逻辑"""
    experiment = get_experiment(exp_id)
    if not experiment:
        return {"success": False, "detail": "实验不存在"}

    existing = [r for r in experiment["readings"] if r["field_key"] == field_key]
    run_index = len(existing) + 1

    mock_enabled = get_config("mock_camera_enabled", default=False)
    try:
        image_dir = get_config("image_dir", default=None) or None
        if mock_enabled:
            client = MockCameraClient(camera_id=camera_id, image_dir=image_dir)
            success, result = client.trigger_and_read()
        else:
            camera_config = Config.get_camera_config()
            if image_dir:
                camera_config["image_dir"] = image_dir
            client = CameraClient(camera_id=camera_id, config=camera_config)
            success, result = client.trigger_and_read()
    except Exception as e:
        logger.error(f"相机 {camera_id} 拍照失败: {e}")
        return {"success": False, "detail": f"相机连接失败: {e}"}

    if not success:
        return {"success": False, "detail": result.get("error", "OCR 识别失败")}

    raw_reading = result.get("reading", "")
    try:
        value = float(str(raw_reading).strip())
    except (ValueError, TypeError):
        logger.error(f"相机 {camera_id} OCR 结果无法解析为数字: {raw_reading!r}")
        return {"success": False, "detail": f"OCR 结果无法解析: {raw_reading}"}

    saved_image_path = None
    # 核心修复：优先使用原始高分辨率路径
    original_path = result.get("raw_image_path") or result.get("image_path")
    if original_path:
        saved_image_path = _convert_and_save_image(original_path, camera_id, run_index)

    reading = create_reading(
        experiment_id=exp_id,
        field_key=field_key,
        camera_id=camera_id,
        value=value,
        run_index=run_index,
        confidence=result.get("confidence"),
        image_path=saved_image_path,
    )
    return {"success": True, "reading": reading}


@app.post("/experiments/{exp_id}/run-async")
def run_experiment_field_async(exp_id: int, body: ExperimentRunField):
    """
    [异步版] 触发单个字段的相机拍照→OCR→保存读数。

    立即返回 task_id，前端轮询 GET /tasks/{task_id} 获取结果。
    """
    experiment = get_experiment(exp_id)
    if not experiment:
        raise HTTPException(status_code=404, detail="实验不存在")

    task_id = task_manager.submit(
        _do_run_experiment_field, exp_id, body.field_key, body.camera_id,
    )
    return {"task_id": task_id, "status": "pending", "message": "任务已提交，请轮询 /tasks/{task_id} 获取结果"}


def _do_run_test_capture(exp_id: int, body_dict: dict):
    """后台线程：执行 run_test_capture 的实际逻辑（委托给 _core_run_test_capture）"""
    return _core_run_test_capture(
        exp_id=exp_id,
        field_key=body_dict["field_key"],
        camera_id=body_dict["camera_id"],
        target_instrument_id=body_dict.get("target_instrument_id"),
        image_path=body_dict.get("image_path"),
        reading_key=body_dict.get("reading_key"),
        run_index=body_dict.get("run_index"),
        precise=body_dict.get("precise", False),
        camera_mode=body_dict.get("camera_mode"),
    )


@app.post("/experiments/{exp_id}/run-test-async")
def run_test_capture_async(exp_id: int, body: ExperimentRunTestBody):
    """
    [异步版] 测试模板 OCR 识别。

    立即返回 task_id，前端轮询 GET /tasks/{task_id} 获取结果。
    """
    experiment = get_experiment(exp_id)
    if not experiment:
        raise HTTPException(status_code=404, detail="实验不存在")

    task_id = task_manager.submit(
        _do_run_test_capture, exp_id, body.model_dump(),
    )
    return {"task_id": task_id, "status": "pending", "message": "任务已提交，请轮询 /tasks/{task_id} 获取结果"}


def _do_read_multi_instruments(body_dict: dict):
    """后台线程：执行 read_multi_instruments 的实际逻辑"""
    from backend.services.multi_instrument_pipeline import MultiInstrumentPipeline
    from PIL import Image

    full_image_path = body_dict.get("image_path")

    if not full_image_path and body_dict.get("camera_id") is not None:
        camera_id = body_dict["camera_id"]
        mock_enabled = get_config("mock_camera_enabled", default=False)
        image_dir = get_config("image_dir", default=None) or None
        try:
            if mock_enabled:
                from backend.services.mock_camera import MockCameraClient
                client = MockCameraClient(camera_id=camera_id, image_dir=image_dir)
                success, result = client.capture_image()
            else:
                from backend.services.camera_control import CameraClient
                camera_config = Config.get_camera_config()
                if image_dir:
                    camera_config["image_dir"] = image_dir
                client = CameraClient(camera_id=camera_id, config=camera_config)
                success, result = client.capture_image()
        except Exception as e:
            logger.error(f"Camera {camera_id} capture failed: {e}")
            return {"success": False, "detections": [], "detail": f"Capture failed: {str(e)}"}

        if not success:
            return {"success": False, "detections": [], "detail": result.get("error", "Capture failed")}
        full_image_path = result.get("image_path")
        if not full_image_path:
            return {"success": False, "detections": [], "detail": "No image captured"}

    try:
        assert full_image_path is not None
        image = Image.open(full_image_path).convert("RGB")
    except Exception as e:
        return {"success": False, "detections": [], "detail": f"Cannot open image: {str(e)}"}

    try:
        pipeline = MultiInstrumentPipeline()
        detections = pipeline.process_image(image)
        return {"success": True, "detections": detections}
    except Exception as e:
        logger.error(f"Pipeline processing failed: {e}")
        return {"success": False, "detections": [], "detail": f"Processing failed: {str(e)}"}


class ReadMultiRequest(BaseModel):
    image_path: Optional[str] = None
    camera_id: Optional[int] = None


@app.post("/read-multi-async")
def read_multi_instruments_async(body: ReadMultiRequest):
    """
    [异步版] Multi-instrument reading endpoint.

    立即返回 task_id，前端轮询 GET /tasks/{task_id} 获取结果。
    """
    task_id = task_manager.submit(
        _do_read_multi_instruments, body.model_dump(),
    )
    return {"task_id": task_id, "status": "pending", "message": "任务已提交，请轮询 /tasks/{task_id} 获取结果"}


@app.get("/config/camera-instruments")
def get_camera_instruments():
    """返回相机编号到仪器的映射（动态从数据库获取）"""
    from instrument_reader import DynamicInstrumentLibrary
    
    # 获取默认路由表 (D0 -> 0, D1 -> 3, etc.)
    route_map = DynamicInstrumentLibrary.get_route_map()
    
    results = {}
    for slot_id, camera_id in route_map.items():
        slot_name = f"D{slot_id}"
        # slot_id 就是 instrument_type 的数字字符串
        template = DynamicInstrumentLibrary.get_template(str(slot_id))
        
        if template:
            try:
                fields_data = json.loads(template['fields_json'])
            except:
                fields_data = []
                
            results[slot_name] = {
                "name": template['name'],
                "readings": [
                    {"key": f["name"], "label": f["label"], "unit": f.get("unit", "")}
                    for f in fields_data
                ]
            }
        else:
            results[slot_name] = {"name": f"原始仪器(D{slot_id})", "readings": []}
            
    return {"success": True, "cameras": results}


class ManualReadingBody(BaseModel):
    field_key: str
    run_index: int
    value: Optional[float] = None
    ocr_data: Optional[Dict[str, Any]] = None
    camera_id: int = 0


@app.put("/experiments/{exp_id}/readings")
def save_manual_reading(exp_id: int, body: ManualReadingBody):
    """手动保存/更新单次读数（用于 OCR 失败时人工填写）"""
    experiment = get_experiment(exp_id)
    if not experiment:
        raise HTTPException(status_code=404, detail="实验不存在")
    reading = upsert_reading(
        experiment_id=exp_id,
        field_key=body.field_key,
        camera_id=body.camera_id,
        value=body.value if body.value is not None else 0.0,
        run_index=body.run_index,
        ocr_data=body.ocr_data
    )
    return {"success": True, "reading": reading}


@app.delete("/experiments/{exp_id}")
def delete_experiment_api(exp_id: int):
    """删除实验记录"""
    deleted = delete_experiment(exp_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="实验不存在")
    return {"success": True, "message": "实验已删除"}


@app.get("/experiments/{exp_id}/export")
def export_experiment(exp_id: int):
    """导出实验数据为 xlsx，按实验类型生成对应模板格式，包含图片路径"""

    experiment = get_experiment(exp_id)
    if not experiment:
        raise HTTPException(status_code=404, detail="实验不存在")

    wb = openpyxl.Workbook()
    ws = wb.active
    exp_type = experiment.get("type", "")
    readings = experiment.get("readings", [])
    manual = experiment.get("manual_params", {})

    # 样式辅助
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    subheader_fill = PatternFill("solid", fgColor="EEF2FA")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(row, col, value, bold=False, fill=None, align="left"):
        c = ws.cell(row=row, column=col, value=value)
        if bold:
            c.font = Font(bold=True)
        if fill:
            c.fill = fill
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        c.border = border
        return c

    def header_row(row, cols_values: list, fill=None):
        """写一行标题，cols_values = [(col, value), ...]"""
        for col, val in cols_values:
            cell(row, col, val, bold=True, fill=fill or header_fill, align="center")

    ws.title = experiment["name"][:31]

    if exp_type == "kinematic_viscosity":
        # ── 标题行
        ws.merge_cells("A1:F1")
        cell(1, 1, "运动粘度检验原始记录", bold=True, align="center")
        ws.cell(1, 1).font = Font(bold=True, size=13)

        # ── 文件信息
        cell(2, 1, "文件编号 File：WLD-QP5100113-02　版次 Edition：A/1", align="left")
        ws.merge_cells("A2:F2")

        # ── 头信息
        header_row(3, [(1, "检测日期"), (3, "报告编号 NO.")], fill=subheader_fill)
        cell(3, 2, manual.get("test_date", ""))
        ws.merge_cells("B3:C3") if False else None  # skip merge for simplicity
        cell(3, 4, manual.get("report_number", ""))
        ws.merge_cells("D3:F3") if False else None

        header_row(4, [(1, "被检样品名称"), (3, "样品编号")], fill=subheader_fill)
        cell(4, 2, manual.get("sample_name", ""))
        cell(4, 4, manual.get("sample_number", ""))

        cell(5, 1, "配方说明", bold=True, fill=subheader_fill)
        cell(5, 2, manual.get("formula_description", ""))
        ws.merge_cells("B5:F5")

        # ── 实验参数
        header_row(6, [(1, "温控设置温度 (℃)"), (2, "最高温度 (℃)"), (3, "最低温度 (℃)"),
                       (4, "毛细管粘度计型号"), (5, "毛细管系数 C (mm²/s²)")], fill=subheader_fill)
        cell(7, 1, manual.get("temperature_set", ""), align="center")
        cell(7, 2, manual.get("temperature_max", ""), align="center")
        cell(7, 3, manual.get("temperature_min", ""), align="center")
        cell(7, 4, manual.get("capillary_model", ""), align="center")
        cell(7, 5, manual.get("capillary_coeff", ""), align="center")

        # ── 流经时间读数
        header_row(9, [(1, "实验次数"), (2, "流经时间 t (s)"), (3, "时间戳"), (4, "图片路径")],
                   fill=header_fill)
        ft_readings = [r for r in readings if r["field_key"] == "flow_time"]
        for i, r in enumerate(ft_readings):
            row_n = 10 + i
            cell(row_n, 1, f"实验 {i + 1}", align="center")
            cell(row_n, 2, r["value"], align="center")
            cell(row_n, 3, (r.get("timestamp") or "")[:19])
            cell(row_n, 4, r.get("image_path") or "")

        # ── 计算结果
        res_row = 10 + len(ft_readings) + 1
        if ft_readings:
            avg = sum(r["value"] for r in ft_readings) / len(ft_readings)
            coeff = float(manual.get("capillary_coeff") or 0)
            cell(res_row, 1, "平均流经时间 t̄ (s)", bold=True, fill=subheader_fill)
            cell(res_row, 2, round(avg, 4), align="center")
            cell(res_row + 1, 1, "运动粘度 ν = C × t̄  (mm²/s)", bold=True, fill=subheader_fill)
            cell(res_row + 1, 2, round(coeff * avg, 4), align="center")

        # 列宽
        for col, width in [(1, 22), (2, 16), (3, 20), (4, 36), (5, 22)]:
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 24

    elif exp_type == "apparent_viscosity":
        ws.merge_cells("A1:F1")
        cell(1, 1, "表观黏度检测原始记录", bold=True, align="center")
        ws.cell(1, 1).font = Font(bold=True, size=13)

        cell(2, 1, "文件编号 File：WLD/CNAS-QP7080004　版次 Edition：A/3")
        ws.merge_cells("A2:F2")

        header_row(3, [(1, "检测日期"), (3, "编号 NO.")], fill=subheader_fill)
        cell(3, 2, manual.get("test_date", ""))
        cell(3, 4, manual.get("report_number", ""))
        header_row(4, [(1, "被检样品名称"), (3, "样品编号")], fill=subheader_fill)
        cell(4, 2, manual.get("sample_name", ""))
        cell(4, 4, manual.get("sample_number", ""))
        cell(5, 1, "配方说明", bold=True, fill=subheader_fill)
        cell(5, 2, manual.get("formula_description", ""))
        ws.merge_cells("B5:F5")

        # ── 表头
        header_row(7, [(1, "实验次数"), (2, "3 rpm"), (3, "6 rpm"), (4, "100 rpm (α)"),
                       (5, "表观黏度 η (mPa·s)"), (6, "图片路径（100rpm）")], fill=header_fill)

        for run_idx in [0, 1]:
            row_n = 8 + run_idx
            rpm3 = next((r["value"] for r in readings if r["field_key"] == "reading_3rpm" and r["run_index"] == run_idx), "")
            rpm6 = next((r["value"] for r in readings if r["field_key"] == "reading_6rpm" and r["run_index"] == run_idx), "")
            rpm100_r = next((r for r in readings if r["field_key"] == "reading_100rpm" and r["run_index"] == run_idx), None)
            rpm100 = rpm100_r["value"] if rpm100_r else ""
            eta = round((rpm100 * 5.077) / 1.704, 4) if rpm100 != "" else ""
            img = rpm100_r.get("image_path") or "" if rpm100_r else ""
            cell(row_n, 1, f"实验 {run_idx + 1}", align="center")
            cell(row_n, 2, rpm3, align="center")
            cell(row_n, 3, rpm6, align="center")
            cell(row_n, 4, rpm100, align="center")
            cell(row_n, 5, eta, align="center")
            cell(row_n, 6, img)

        all_eta = [(r["value"] * 5.077) / 1.704 for r in readings if r["field_key"] == "reading_100rpm"]
        if all_eta:
            cell(11, 1, "平均表观黏度 (mPa·s)", bold=True, fill=subheader_fill)
            cell(11, 5, round(sum(all_eta) / len(all_eta), 4), align="center")

        # ── 每个字段的完整读数+图片
        row_n = 13
        cell(row_n, 1, "各读数详情（含图片路径）", bold=True, fill=subheader_fill)
        ws.merge_cells(f"A{row_n}:F{row_n}")
        row_n += 1
        header_row(row_n, [(1, "字段"), (2, "实验次数"), (3, "读数值"), (4, "时间戳"), (5, "图片路径")],
                   fill=header_fill)
        row_n += 1
        for r in readings:
            cell(row_n, 1, r["field_key"])
            cell(row_n, 2, f"实验 {r['run_index'] + 1}", align="center")
            cell(row_n, 3, r["value"], align="center")
            cell(row_n, 4, (r.get("timestamp") or "")[:19])
            cell(row_n, 5, r.get("image_path") or "")
            row_n += 1

        for col, width in [(1, 18), (2, 10), (3, 12), (4, 12), (5, 20), (6, 36)]:
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 24

    elif exp_type == "surface_tension":
        ws.merge_cells("A1:F1")
        cell(1, 1, "表面张力和界面张力检测原始记录", bold=True, align="center")
        ws.cell(1, 1).font = Font(bold=True, size=13)

        cell(2, 1, "文件编号 File：WLD/CNAS-QP7080004　版次 Edition：A/3　执行标准：SY/T 5370-2018")
        ws.merge_cells("A2:F2")

        # ── 头信息
        for r_idx, (lbl1, key1, lbl2, key2) in enumerate([
            ("检测日期", "test_date", "检测报告编号", "report_number"),
            ("被检样品名称", "sample_name", "样品编号", "sample_number"),
            ("样品状态", "sample_state", "配液编号", "formula_number"),
        ], start=3):
            cell(r_idx, 1, lbl1, bold=True, fill=subheader_fill)
            cell(r_idx, 2, manual.get(key1, ""))
            cell(r_idx, 3, lbl2, bold=True, fill=subheader_fill)
            cell(r_idx, 4, manual.get(key2, ""))

        cell(6, 1, "配方说明", bold=True, fill=subheader_fill)
        cell(6, 2, manual.get("formula_description", ""))
        ws.merge_cells("B6:F6")

        # ── 环境条件
        header_row(7, [(1, "室内温度 (℃)"), (2, "室内湿度 (%)"),
                       (3, "25℃ 破胶液密度 (g/cm³)"), (4, "25℃ 煤油密度 (g/cm³)")],
                   fill=subheader_fill)
        cell(8, 1, manual.get("room_temperature", ""), align="center")
        cell(8, 2, manual.get("room_humidity", ""), align="center")
        cell(8, 3, manual.get("sample_density", ""), align="center")
        cell(8, 4, manual.get("kerosene_density", ""), align="center")

        # ── 纯水表面张力
        cell(10, 1, "纯水表面张力验证", bold=True, fill=subheader_fill)
        ws.merge_cells("A10:F10")
        header_row(11, [(1, "实验次数"), (2, "测试值 (mN/m)"), (3, "时间戳"), (4, "图片路径")],
                   fill=header_fill)
        wst = [r for r in readings if r["field_key"] == "water_surface_tension"]
        for i, r in enumerate(wst):
            rn = 12 + i
            cell(rn, 1, f"实验 {i + 1}", align="center")
            cell(rn, 2, r["value"], align="center")
            cell(rn, 3, (r.get("timestamp") or "")[:19])
            cell(rn, 4, r.get("image_path") or "")
        wst_avg_row = 12 + len(wst)
        if wst:
            cell(wst_avg_row, 1, "算术平均值", bold=True, fill=subheader_fill)
            cell(wst_avg_row, 2, round(sum(r["value"] for r in wst) / len(wst), 4), align="center")

        # ── 破胶液表面张力
        base_fst = wst_avg_row + 2
        cell(base_fst, 1, f"破胶液表面张力（样品密度 {manual.get('sample_density','')} g/cm³）",
             bold=True, fill=subheader_fill)
        ws.merge_cells(f"A{base_fst}:F{base_fst}")
        header_row(base_fst + 1, [(1, "实验次数"), (2, "表面张力 (mN/m)"), (3, "时间戳"), (4, "图片路径")],
                   fill=header_fill)
        fst = [r for r in readings if r["field_key"] == "fluid_surface_tension"]
        for i, r in enumerate(fst):
            rn = base_fst + 2 + i
            cell(rn, 1, f"实验 {i + 1}", align="center")
            cell(rn, 2, r["value"], align="center")
            cell(rn, 3, (r.get("timestamp") or "")[:19])
            cell(rn, 4, r.get("image_path") or "")
        avg_row = base_fst + 2 + len(fst)
        if fst:
            cell(avg_row, 1, "算术平均值", bold=True, fill=subheader_fill)
            cell(avg_row, 2, round(sum(r["value"] for r in fst) / len(fst), 4), align="center")

        # ── 破胶液界面张力
        base = avg_row + 2
        cell(base, 1, f"破胶液界面张力（煤油密度 {manual.get('kerosene_density','')} g/cm³）",
             bold=True, fill=subheader_fill)
        ws.merge_cells(f"A{base}:F{base}")
        header_row(base + 1, [(1, "实验次数"), (2, "界面张力 (mN/m)"), (3, "时间戳"), (4, "图片路径")],
                   fill=header_fill)
        fit = [r for r in readings if r["field_key"] == "fluid_interface_tension"]
        for i, r in enumerate(fit):
            rn = base + 2 + i
            cell(rn, 1, f"实验 {i + 1}", align="center")
            cell(rn, 2, r["value"], align="center")
            cell(rn, 3, (r.get("timestamp") or "")[:19])
            cell(rn, 4, r.get("image_path") or "")
        if fit:
            avg_r = base + 2 + len(fit)
            cell(avg_r, 1, "算术平均值", bold=True, fill=subheader_fill)
            cell(avg_r, 2, round(sum(r["value"] for r in fit) / len(fit), 4), align="center")
            sign_row = avg_r + 2
        else:
            sign_row = base + 4

        # ── 签署区
        cell(sign_row, 1, "备注", bold=True, fill=subheader_fill)
        cell(sign_row, 2, manual.get("remarks", ""))
        ws.merge_cells(f"B{sign_row}:F{sign_row}")
        cell(sign_row + 1, 1, "检测人", bold=True, fill=subheader_fill)
        cell(sign_row + 1, 2, manual.get("operator_name", ""))
        cell(sign_row + 1, 3, "检测时间", bold=True, fill=subheader_fill)
        cell(sign_row + 1, 4, manual.get("operator_time", ""))
        cell(sign_row + 2, 1, "审核人", bold=True, fill=subheader_fill)
        cell(sign_row + 2, 2, manual.get("reviewer_name", ""))
        cell(sign_row + 2, 3, "审核日期", bold=True, fill=subheader_fill)
        cell(sign_row + 2, 4, manual.get("reviewer_date", ""))

        for col, width in [(1, 22), (2, 16), (3, 20), (4, 36)]:
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 24

    elif exp_type == "ph_value":
        ws.merge_cells("A1:F1")
        cell(1, 1, "pH值检测原始记录", bold=True, align="center")
        ws.cell(1, 1).font = Font(bold=True, size=13)

        cell(2, 1, "文件编号 File：WLD-QP5100113-02　版次 Edition：A/1　执行标准：SY/T 5107-2016")
        ws.merge_cells("A2:F2")

        header_row(3, [(1, "检测日期"), (3, "样品编号")], fill=subheader_fill)
        cell(3, 2, manual.get("test_date", ""))
        cell(3, 4, manual.get("sample_number", ""))
        header_row(4, [(1, "被检样品名称"), (3, "室内温度 (℃)"), (5, "室内湿度 (%)")], fill=subheader_fill)
        cell(4, 2, manual.get("sample_name", ""))
        cell(4, 4, manual.get("room_temp", ""))
        cell(4, 6, manual.get("room_humidity", ""))

        header_row(6, [(1, "实验次数"), (2, "pH值"), (3, "温度 (℃)"), (4, "时间戳"), (5, "图片路径")], fill=header_fill)
        
        ph_readings = [r for r in readings if r["field_key"] == "ph_measurement"]
        ph_vals = []
        temp_vals = []
        
        for i, r in enumerate(ph_readings):
            rn = 7 + i
            ocr = r.get("ocr_data", {})
            ph = ocr.get("ph_value") or r.get("value")
            temp = ocr.get("temperature")
            
            if ph is not None and not isinstance(ph, str): ph_vals.append(float(ph))
            elif ph and str(ph).replace('.', '', 1).isdigit(): ph_vals.append(float(ph))
                
            if temp is not None and not isinstance(temp, str): temp_vals.append(float(temp))
            elif temp and str(temp).replace('.', '', 1).isdigit(): temp_vals.append(float(temp))
                
            cell(rn, 1, f"实验 {i + 1}", align="center")
            cell(rn, 2, ph if ph is not None else "", align="center")
            cell(rn, 3, temp if temp is not None else "", align="center")
            cell(rn, 4, (r.get("timestamp") or "")[:19])
            cell(rn, 5, r.get("image_path") or "")
            
        avg_row = 7 + len(ph_readings)
        if ph_readings:
            cell(avg_row, 1, "算术平均值", bold=True, fill=subheader_fill)
            cell(avg_row, 2, round(sum(ph_vals) / len(ph_vals), 2) if ph_vals else "", align="center")
            cell(avg_row, 3, round(sum(temp_vals) / len(temp_vals), 1) if temp_vals else "", align="center")
            
        for col, width in [(1, 12), (2, 12), (3, 12), (4, 20), (5, 36)]:
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 24

    elif exp_type == "test":
        ws.merge_cells("A1:F1")
        cell(1, 1, "全相机测试拍照记录", bold=True, align="center")
        ws.cell(1, 1).font = Font(bold=True, size=13)
        cell(2, 1, f"实验名称: {experiment['name']}")
        ws.merge_cells("A2:F2")

        header_row(4, [(1, "相机位"), (2, "字段"), (3, "序号"), (4, "读数值"),
                       (5, "时间戳"), (6, "图片路径")], fill=header_fill)
        row_n = 5
        for cam_id in range(9):
            cam_readings = [r for r in readings if r["camera_id"] == cam_id]
            for i, r in enumerate(cam_readings):
                cell(row_n, 1, f"F{cam_id}", align="center")
                cell(row_n, 2, r["field_key"])
                cell(row_n, 3, i + 1, align="center")
                cell(row_n, 4, r["value"], align="center")
                cell(row_n, 5, (r.get("timestamp") or "")[:19])
                cell(row_n, 6, r.get("image_path") or "")
                row_n += 1

        for col, width in [(1, 10), (2, 16), (3, 8), (4, 12), (5, 20), (6, 36)]:
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"experiment_{exp_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ==================== 系统配置 API ====================

@app.get("/config/mock")
def get_mock_config():
    """获取 Mock 相机模式开关状态"""
    enabled = get_config("mock_camera_enabled", default=False)
    return {"mock_enabled": bool(enabled)}


class MockConfigUpdate(BaseModel):
    enabled: bool


@app.post("/config/mock")
def set_mock_config(body: MockConfigUpdate):
    """设置 Mock 相机模式开关"""
    set_config("mock_camera_enabled", body.enabled)
    logger.info(f"Mock 相机模式已{'开启' if body.enabled else '关闭'}")
    return {"mock_enabled": body.enabled, "message": f"Mock 模式已{'开启' if body.enabled else '关闭'}"}


@app.get("/config/image-dir")
def get_image_dir():
    """获取图片存储目录配置"""
    image_dir = get_config("image_dir", default="")
    return {"image_dir": image_dir}


class ImageDirUpdate(BaseModel):
    image_dir: str


@app.post("/config/image-dir")
def set_image_dir_config(body: ImageDirUpdate):
    """设置图片存储目录"""
    set_config("image_dir", body.image_dir)
    logger.info(f"图片存储目录已设置为: {body.image_dir}")
    return {"image_dir": body.image_dir, "message": "图片存储目录已更新"}


# ==================== 工具 API ====================

@app.get("/templates")
def list_templates():
    """获取所有仪器模板"""
    templates = get_all_templates()
    for t in templates:
        t['fields'] = json.loads(t['fields_json'])
        t['keywords'] = json.loads(t['keywords_json'])
        t['example_images'] = json.loads(t['example_images_json'] or '[]')
    return {"success": True, "templates": templates}

@app.post("/templates")
def create_or_update_template(body: InstrumentTemplateCreate):
    """创建或更新仪器模板"""
    upsert_template(
        instrument_type=body.instrument_type,
        name=body.name,
        description=body.description,
        prompt_template=body.prompt_template,
        fields=[f.model_dump() for f in body.fields],
        keywords=body.keywords,
        example_images=body.example_images,
        default_tier=body.default_tier
    )
    return {"success": True, "message": "Template saved"}

# ==================== LLM 模型配置 API ====================

def _mask_api_key(key: Optional[str]) -> Optional[str]:
    """脱敏 API Key，仅显示前4位和后4位"""
    if not key or len(key) <= 8:
        return key
    return key[:4] + "****" + key[-4:]


@app.get("/config/llm")
def get_llm_config():
    """获取当前 LLM 模型配置（API Key 脱敏）"""
    saved = get_config("llm_config", default=None)
    if saved:
        saved = {**saved, "api_key": _mask_api_key(saved.get("api_key"))}
        return {"success": True, "config": saved}
    return {
        "success": True,
        "config": {
            "provider": "local_vlm",
            "model_name": "GLM-OCR",
            "base_url": Config.LOCAL_VLM_PATH,
            "api_key": None,
            "temperature": Config.MODEL_TEMPERATURE,
            "max_tokens": Config.MODEL_MAX_TOKENS,
        }
    }


class LLMConfigUpdate(BaseModel):
    provider: str
    model_name: str
    base_url: str
    api_key: Optional[str] = None
    temperature: Optional[float] = None
    max_tokens: Optional[int] = None


@app.post("/config/llm")
def set_llm_config(body: LLMConfigUpdate):
    """设置 LLM 模型配置并重建 provider"""
    from backend.services.llm_provider import create_provider, LLMConfig, set_global_provider

    if body.provider not in ("openai_compatible", "local_vlm"):
        raise HTTPException(status_code=400, detail=f"不支持的 provider 类型: {body.provider}")

    # 如果前端传回脱敏后的 key（包含 ****），保留数据库中的原始 key
    api_key = body.api_key
    if api_key and "****" in api_key:
        saved = get_config("llm_config", default=None)
        api_key = saved.get("api_key") if isinstance(saved, dict) else None

    config = LLMConfig(
        provider=body.provider,
        model_name=body.model_name,
        base_url=body.base_url,
        api_key=api_key,
        temperature=body.temperature if body.temperature is not None else Config.MODEL_TEMPERATURE,
        max_tokens=body.max_tokens if body.max_tokens is not None else Config.MODEL_MAX_TOKENS,
    )

    try:
        provider = create_provider(config)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法创建 LLM 客户端: {str(e)}")

    config_dict = {
        "provider": config.provider,
        "model_name": config.model_name,
        "base_url": config.base_url,
        "api_key": config.api_key,
        "temperature": config.temperature,
        "max_tokens": config.max_tokens,
    }
    set_config("llm_config", config_dict)
    set_global_provider(provider)

    # 返回脱敏后的配置
    response_config = {**config_dict, "api_key": _mask_api_key(config.api_key)}
    logger.info(f"LLM 模型已切换: {config.provider}/{config.model_name} @ {config.base_url}")
    return {"success": True, "config": response_config, "message": "模型配置已更新"}


@app.get("/config/llm/models")
def list_llm_models(base_url: Optional[str] = None):
    """列出 LMStudio 服务上的可用模型"""
    if base_url:
        url = base_url
    else:
        saved = get_config("llm_config", default=None)
        if isinstance(saved, dict):
            url = saved.get("base_url", Config.LMSTUDIO_BASE_URL)
        else:
            url = Config.LMSTUDIO_BASE_URL
    try:
        import httpx
        resp = httpx.get(f"{url}/api/tags", timeout=5.0)
        resp.raise_for_status()
        models = resp.json().get("models", [])
        return {
            "success": True,
            "models": [
                {"name": m["name"], "size": m.get("size", 0), "modified_at": m.get("modified_at", "")}
                for m in models
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"无法连接 LMStudio 服务: {str(e)}")


@app.get("/config/llm/status")
def check_llm_status():
    """检查 LLM 模型连接状态（使用轻量级 API，不消耗 GPU）"""
    try:
        from backend.services.llm_provider import get_global_provider
        provider = get_global_provider()
        if provider.provider_type == "local_vlm":
            import torch
            gpu_info = None
            if torch.cuda.is_available():
                gpu_info = {
                    "name": torch.cuda.get_device_name(0),
                    "memory_allocated": f"{torch.cuda.memory_allocated(0) / 1024**2:.2f} MB",
                    "memory_reserved": f"{torch.cuda.memory_reserved(0) / 1024**2:.2f} MB",
                }
            return {
                "success": True,
                "status": "ready",
                "provider": provider.provider_type,
                "model": provider.model_name,
                "gpu": gpu_info
            }
            
        return {
            "success": True,
            "status": "connected",
            "provider": provider.provider_type,
            "model": provider.model_name,
        }
    except Exception as e:
        return {
            "success": False,
            "status": "error",
            "detail": str(e),
        }

@app.get("/health")
def health_check():
    """健康检查"""
    return {
        "success": True,
        "timestamp": datetime.now().isoformat(),
        "message": "OCR 实验服务运行中"
    }


@app.get("/config")
def get_system_config():
    """获取系统配置"""
    return {
        "success": True,
        "config": {
            "camera_count": Config.CAMERA_COUNT,
            "camera_control_host": Config.CAMERA_CONTROL_HOST,
            "camera_control_port": Config.CAMERA_CONTROL_PORT,
            "lmstudio_base_url": Config.LMSTUDIO_BASE_URL,
            "lmstudio_model": Config.LMSTUDIO_MODEL
        }
    }


# ==================== Multi-Instrument Pipeline API ====================

@app.post("/read-multi")
def read_multi_instruments(body: ReadMultiRequest):
    """Multi-instrument reading endpoint.
    Detects, classifies, and reads all instruments in the image.
    """
    from backend.services.multi_instrument_pipeline import MultiInstrumentPipeline
    from PIL import Image

    full_image_path = body.image_path

    if not full_image_path and body.camera_id is not None:
        mock_enabled = get_config("mock_camera_enabled", default=False)
        image_dir = get_config("image_dir", default=None) or None
        try:
            if mock_enabled:
                from backend.services.mock_camera import MockCameraClient
                client = MockCameraClient(camera_id=body.camera_id, image_dir=image_dir)
                success, result = client.capture_image()
            else:
                from backend.services.camera_control import CameraClient
                camera_config = Config.get_camera_config()
                if image_dir:
                    camera_config["image_dir"] = image_dir
                client = CameraClient(camera_id=body.camera_id, config=camera_config)
                success, result = client.capture_image()
        except Exception as e:
            logger.error(f"Camera {body.camera_id} capture failed: {e}")
            return {"success": False, "detections": [], "detail": f"Capture failed: {str(e)}"}

        if not success:
            return {"success": False, "detections": [], "detail": result.get("error", "Capture failed")}
        full_image_path = result.get("image_path")
        if not full_image_path:
            return {"success": False, "detections": [], "detail": "No image captured"}

    try:
        assert full_image_path is not None
        image = Image.open(full_image_path).convert("RGB")
    except Exception as e:
        return {"success": False, "detections": [], "detail": f"Cannot open image: {str(e)}"}

    try:
        pipeline = MultiInstrumentPipeline()
        detections = pipeline.process_image(image)
        return {"success": True, "detections": detections}
    except Exception as e:
        logger.error(f"Pipeline processing failed: {e}")
        return {"success": False, "detections": [], "detail": f"Processing failed: {str(e)}"}


@app.post("/rebuild-clip-cache")
def rebuild_clip_cache():
    """Rebuild CLIP embedding cache after template/reference image changes"""
    from backend.services.clip_matcher import CLIPInstrumentMatcher
    try:
        matcher = CLIPInstrumentMatcher()
        matcher.invalidate_cache()
        return {"success": True, "message": "CLIP cache rebuilt successfully"}
    except Exception as e:
        logger.error(f"Failed to rebuild CLIP cache: {e}")
        return {"success": False, "detail": str(e)}


# 仪器模板与映射管理
@app.get("/instruments/templates")
def list_instrument_templates():
    """获取所有仪器模板（仅返回 D0-D8 核心 9 个仪器）"""
    templates = get_all_templates()
    # 为前端格式化：仅返回 D0-D8
    result = []
    # 核心仪器 ID 列表 0-8
    core_ids = [str(i) for i in range(9)]
    
    for t in templates:
        inst_type = t["instrument_type"]
        if inst_type in core_ids or (inst_type.startswith('D') and inst_type[1:] in core_ids) or (inst_type.startswith('F') and inst_type[1:] in core_ids):
            # 统一展示为 D0, D1... 格式
            id_str = inst_type[1:] if inst_type.startswith(('D', 'F')) else inst_type
            display_type = f"D{id_str}"
            result.append({
                "instrument_type": display_type,
                "name": t["name"],
                "description": t["description"]
            })
    
    # 按 D0-D8 排序
    result.sort(key=lambda x: int(x["instrument_type"][1:]))
    return {"success": True, "templates": result}


@app.get("/config/instrument-camera-mapping")
def get_instrument_mapping():
    """获取当前的仪器-相机映射配置"""
    mapping = get_config("instrument_camera_mapping", default={})
    return {"success": True, "mapping": mapping}


@app.post("/config/instrument-camera-mapping")
def update_instrument_mapping(body: InstrumentMappingUpdate):
    """手动更新仪器-相机映射配置"""
    try:
        # TODO: 校验 instrument_id 是否在模板中，camera_id 是否存在
        set_config("instrument_camera_mapping", body.mapping)
        logger.info(f"手动更新仪器映射: {body.mapping}")
        return {"success": True, "message": "映射配置已更新"}
    except Exception as e:
        logger.error(f"更新仪器映射失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
